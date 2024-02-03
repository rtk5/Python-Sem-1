import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import Workbook, load_workbook
import os

class UserProfile:
    def __init__(self, username, display_name, budget_limit, profile_picture):
        self.username = username
        self.display_name = display_name
        self.budget_limit = budget_limit
        self.profile_picture = profile_picture

class UserProfileApp:
    def __init__(self, root):
        self.root = root
        self.root.title("User Profile")

        # Widgets for user input
        self.label_username = tk.Label(root, text="Username:")
        self.entry_username = tk.Entry(root)

        self.label_display_name = tk.Label(root, text="Display Name:")
        self.entry_display_name = tk.Entry(root)

        self.label_budget_limit = tk.Label(root, text="Budget Limit:")
        self.entry_budget_limit = tk.Entry(root)

        self.label_profile_picture = tk.Label(root, text="Profile Picture:")
        self.entry_profile_picture = tk.Entry(root)
        self.button_browse = tk.Button(root, text="Browse", command=self.browse_profile_picture)

        # Widget for creating the user profile
        self.button_create_profile = tk.Button(root, text="Create Profile", command=self.create_user_profile)

        # Arrange widgets
        self.label_username.grid(row=0, column=0, padx=5, pady=5)
        self.entry_username.grid(row=0, column=1, padx=5, pady=5)

        self.label_display_name.grid(row=1, column=0, padx=5, pady=5)
        self.entry_display_name.grid(row=1, column=1, padx=5, pady=5)

        self.label_budget_limit.grid(row=2, column=0, padx=5, pady=5)
        self.entry_budget_limit.grid(row=2, column=1, padx=5, pady=5)

        self.label_profile_picture.grid(row=3, column=0, padx=5, pady=5)
        self.entry_profile_picture.grid(row=3, column=1, padx=5, pady=5)
        self.button_browse.grid(row=3, column=2, padx=5, pady=5)

        self.button_create_profile.grid(row=4, column=0, columnspan=2, pady=10)

    def browse_profile_picture(self):
        file_path = filedialog.askopenfilename(title="Select Profile Picture", filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
        self.entry_profile_picture.delete(0, tk.END)
        self.entry_profile_picture.insert(0, file_path)

    def create_user_profile(self):
        username = self.entry_username.get()
        display_name = self.entry_display_name.get()
        budget_limit = self.entry_budget_limit.get()
        profile_picture = self.entry_profile_picture.get()

        user = UserProfile(username=username, display_name=display_name, budget_limit=budget_limit, profile_picture=profile_picture)
        self.save_to_excel(user)

    def save_to_excel(self, user):
        excel_file_path = r"C:\Users\rithv\WORK\PES\Miniproject\budget_tracker1.xlsx"
        try:
            if os.path.exists(excel_file_path):
                workbook = load_workbook(excel_file_path)
            else:
                workbook = Workbook()
                workbook.remove(workbook.active)
                workbook.save(excel_file_path)

            sheet = workbook.create_sheet(title=user.username)
            sheet.append(["Username", "Display Name", "Budget Limit", "Profile Picture"])
            sheet.append([user.username, user.display_name, user.budget_limit, user.profile_picture])

            workbook.save(excel_file_path)
            messagebox.showinfo("User Profile", "User profile saved successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

if __name__ == "__main__":
    root = tk.Tk()
    app = UserProfileApp(root)
    root.mainloop()

